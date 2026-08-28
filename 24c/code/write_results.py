# -*- coding: utf-8 -*-
"""将求解方案写入 result 模板 xlsx（严格保持模板格式）。
方案: dict[(yr, season, p)] = crop_id  （crops_data 的作物编号）
"""
import warnings; warnings.filterwarnings('ignore')
import openpyxl, numpy as np, pandas as pd
import crops_data as cd

TEMPLATE = f'{cd.C}/附件3/result1_1.xlsx'

def write_result(plant, out_path, template=TEMPLATE):
    """plant: dict[(yr, season, p)] = cid"""
    wb = openpyxl.load_workbook(template)
    # 建立 (季次,地块名) -> 行号
    for ws in wb.worksheets:
        yr = int(ws.title)
        # 该 sheet 中 (季次,地块) -> 行
        for row in range(1, ws.max_row + 1):
            season = ws.cell(row, 1).value
            plot = ws.cell(row, 2).value
            if season is None or plot is None:
                continue
            season = str(season).replace('\n', '')
            plot = str(plot).strip()
            # 填种植面积
            cid = plant.get((yr, season, plot))
            if cid is not None:
                crop_name = cd.crop_id[cid]
                # 找到作物列
                for col in range(3, ws.max_column + 1):
                    if ws.cell(1, col).value is not None and \
                       str(ws.cell(1, col).value).strip() == crop_name:
                        ws.cell(row, col).value = cd.land_area[plot]
                        break
    wb.save(out_path)
    print(f'已写入 {out_path}')

def write_result_from_plan(plant, out_path):
    write_result(plant, out_path)

if __name__ == '__main__':
    import sys
    write_result_from_plan({}, 'test_empty.xlsx')
